#!/usr/bin/env python3
"""
Enrich emails từ website cho file output của scan_places.py

Usage:
    python3 enrich_emails.py --input leads.xlsx --output leads-enriched.xlsx
"""

import argparse
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import urljoin, urlparse

import requests
from openpyxl import Workbook, load_workbook

EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
EXCLUDE_EMAILS = {"example@", "your@", "email@", "name@", "@sentry.io", "@wixpress.com"}
CONTACT_PATHS = ["", "/contact", "/lien-he", "/about", "/gioi-thieu", "/contact-us"]
HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; HPG-Lead-Enricher/1.0; B2B research)"
}
TIMEOUT = 8


def is_valid_email(email: str) -> bool:
    email_lower = email.lower()
    if any(ex in email_lower for ex in EXCLUDE_EMAILS):
        return False
    if email_lower.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg")):
        return False
    return True


def extract_emails_from_url(url: str) -> list[str]:
    """Try to fetch URL and extract emails."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=TIMEOUT, allow_redirects=True)
        if resp.status_code != 200:
            return []
        emails = EMAIL_REGEX.findall(resp.text)
        return [e for e in emails if is_valid_email(e)]
    except Exception:
        return []


def find_email_for_website(website: str) -> str:
    """Try multiple paths on the website to find an email."""
    if not website or not website.startswith(("http://", "https://")):
        return ""

    parsed = urlparse(website)
    base = f"{parsed.scheme}://{parsed.netloc}"

    found_emails = set()
    for path in CONTACT_PATHS:
        url = urljoin(base, path)
        emails = extract_emails_from_url(url)
        found_emails.update(emails)
        if found_emails:
            break

    if not found_emails:
        return ""

    domain = parsed.netloc.replace("www.", "")
    domain_emails = [e for e in found_emails if domain in e.lower()]
    if domain_emails:
        return domain_emails[0]
    return next(iter(found_emails))


def load_rows(input_path: str) -> tuple[list[str], list[dict]]:
    wb = load_workbook(input_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("Input file is empty")
    headers = list(rows[0])
    data = [dict(zip(headers, row)) for row in rows[1:]]
    return headers, data


def save_rows(headers: list[str], rows: list[dict], output_path: str):
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Places"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    wb.save(output_path)


def main():
    ap = argparse.ArgumentParser(description="Enrich emails từ website")
    ap.add_argument("--input", required=True, help="File Excel input")
    ap.add_argument("--output", required=True, help="File Excel output")
    ap.add_argument("--workers", type=int, default=10, help="Số worker threads")
    args = ap.parse_args()

    headers, rows = load_rows(args.input)
    if "email" not in headers:
        headers.append("email")

    targets = [(i, r) for i, r in enumerate(rows) if r.get("website") and not r.get("email")]
    print(f"Loaded {len(rows)} rows, enriching {len(targets)} with websites...")

    enriched = 0
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        futures = {ex.submit(find_email_for_website, r["website"]): i for i, r in targets}
        for i, fut in enumerate(as_completed(futures), 1):
            idx = futures[fut]
            try:
                email = fut.result()
            except Exception:
                email = ""
            rows[idx]["email"] = email
            if email:
                enriched += 1
                print(f"  [{i}/{len(targets)}] {rows[idx].get('name', '')}: {email}")

    print(f"\nEnriched {enriched}/{len(targets)} emails")
    save_rows(headers, rows, args.output)
    print(f"Saved to {args.output}")


if __name__ == "__main__":
    main()
